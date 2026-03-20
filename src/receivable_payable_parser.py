# -*- coding: utf-8 -*-
"""应收/应付明细解析模块 — 支持 .xls 和 .xlsx，多种格式"""

import pandas as pd
import openpyxl
import xlrd
from typing import List, Tuple, Optional
from datetime import datetime, timedelta
import warnings

warnings.filterwarnings('ignore', category=UserWarning)


def parse(receivable_paths: List[str], payable_paths: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    解析应收和应付明细文件。

    Returns:
        (recv_df, pay_df)
        recv_df 列: ['客户名称', '币种', '预算汇率', '合计(外币)', '合计(万元)', ...各月列]
        pay_df 列:  ['供应商', '合计(万元)', ...各月列]
    """
    recv_dfs = []
    for fpath in receivable_paths:
        try:
            df = _parse_receivable(fpath)
            if df is not None and len(df) > 0:
                recv_dfs.append(df)
        except Exception as e:
            print(f"[WARN] 解析应收文件失败: {fpath}, 错误: {e}")

    pay_dfs = []
    for fpath in payable_paths:
        try:
            df = _parse_payable(fpath)
            if df is not None and len(df) > 0:
                pay_dfs.append(df)
        except Exception as e:
            print(f"[WARN] 解析应付文件失败: {fpath}, 错误: {e}")

    recv_df = pd.concat(recv_dfs, ignore_index=True) if recv_dfs else pd.DataFrame()
    pay_df = pd.concat(pay_dfs, ignore_index=True) if pay_dfs else pd.DataFrame()

    return recv_df, pay_df


def _read_all_rows(fpath: str, sheet_index: int = 0) -> List[list]:
    """读取文件所有行（支持 .xls 和 .xlsx）"""
    if fpath.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws_name = wb.sheetnames[sheet_index] if sheet_index < len(wb.sheetnames) else wb.sheetnames[0]
        ws = wb[ws_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        return data
    elif fpath.lower().endswith('.xls'):
        wb = xlrd.open_workbook(fpath)
        ws = wb.sheet_by_index(min(sheet_index, wb.nsheets - 1))
        data = []
        for row_idx in range(ws.nrows):
            data.append([ws.cell_value(row_idx, col) for col in range(ws.ncols)])
        return data
    return []


def _parse_receivable(fpath: str) -> Optional[pd.DataFrame]:
    """
    解析应收明细账。支持两种格式：
    格式1 (矩阵): 客户名称 + 各月(应收/已收/余额) + 明细账累计
    格式2 (简单): 序号/客户名称/汇率/各月金额/合计
    """
    data = _read_all_rows(fpath)
    if len(data) < 2:
        return None

    # 检测格式：看第1-2行是否有"明细账累计"或"应收金额"
    header_texts = ' '.join([str(c) for c in (data[0] + data[1]) if c])

    if '明细' in header_texts and '累计' in header_texts:
        return _parse_receivable_matrix(data)
    elif '应收金额' in header_texts:
        return _parse_receivable_matrix(data)
    else:
        return _parse_receivable_simple(data)


def _parse_receivable_matrix(data: list) -> Optional[pd.DataFrame]:
    """解析矩阵格式应收（第1种格式）"""
    if len(data) < 4:
        return None

    header_row1 = data[0]
    header_row2 = data[1]

    # 找"明细账累计"列位置
    total_col_start = None
    for i, val in enumerate(header_row1):
        if val and '明细' in str(val) and '累计' in str(val):
            total_col_start = i
            break

    if total_col_start is None:
        total_col_start = len(header_row1) - 3

    results = []
    for row_idx in range(3, len(data)):
        row = data[row_idx]
        if not row or not row[0] or str(row[0]).strip() == '':
            continue

        customer = str(row[0]).strip()
        total_recv = _safe_float(row[total_col_start]) if total_col_start < len(row) else 0
        total_paid = _safe_float(row[total_col_start + 1]) if total_col_start + 1 < len(row) else 0
        balance = _safe_float(row[total_col_start + 2]) if total_col_start + 2 < len(row) else 0

        if total_recv == 0:
            total_recv, total_paid, balance = _manual_sum_months(row, header_row2, total_col_start)

        results.append({
            '客户名称': customer,
            '累计应收': total_recv / 10000,
            '累计已收': total_paid / 10000,
            '应收余额': (total_recv - total_paid) / 10000 if total_recv and total_paid else (balance or 0) / 10000,
        })

    return pd.DataFrame(results)


def _parse_receivable_simple(data: list) -> Optional[pd.DataFrame]:
    """
    解析简单格式应收（如有合计列的客户×月度表）。
    格式: 序号/客户名称/预算汇率/月份1(币别)/月份2(币别)/.../合计(币别)
    自动从列名提取币种信息。
    """
    if len(data) < 2:
        return None

    # 找到包含"客户名称"或类似的表头行
    header_row_idx = 0
    for i, row in enumerate(data[:5]):
        row_text = [str(c).strip() for c in row if c]
        if any('客户' in t or '名称' in t for t in row_text):
            header_row_idx = i
            break

    headers = [str(c).strip() if c else '' for c in data[header_row_idx]]

    # 从列名提取币种 (如 "2025.8月\n币别：USD")
    import re
    currency = 'CNY'
    for h in headers:
        m = re.search(r'币别[：:]\s*(\w+)', h)
        if m:
            currency = m.group(1).upper()
            break

    # 找各关键列
    name_col = None
    rate_col = None
    total_col = None
    month_cols = []  # (col_index, month_label)
    for i, h in enumerate(headers):
        if '客户' in h or '名称' in h:
            name_col = i
        elif '汇率' in h:
            rate_col = i
        elif '合计' in h:
            total_col = i
        elif re.search(r'\d+\.\d+月', h):
            # 提取月份标签 "2025.8月\n币别：USD" -> "2025.08"
            m_month = re.search(r'(\d{4})\.(\d{1,2})月', h)
            if m_month:
                label = f"{m_month.group(1)}-{int(m_month.group(2)):02d}"
                month_cols.append((i, label))

    if name_col is None:
        name_col = 1

    results = []
    for row_idx in range(header_row_idx + 1, len(data)):
        row = data[row_idx]
        if not row:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) else ''
        if not name or name == 'nan' or name in ('', 'None'):
            continue

        # 检查是否为汇总行
        rate_val = str(row[rate_col]).strip() if rate_col and rate_col < len(row) else ''
        is_summary = '小计' in name or '合计' in name or '小计' in rate_val
        if is_summary:
            continue

        # 汇率
        exchange_rate = _safe_float(row[rate_col]) if rate_col and rate_col < len(row) else 0

        # 合计（外币）
        total_foreign = _safe_float(row[total_col]) if total_col and total_col < len(row) else 0

        # 各月金额（外币）
        rec = {
            '客户名称': name,
            '币种': currency,
            '预算汇率': exchange_rate,
            '合计(外币)': total_foreign,
            '合计(万元)': round(total_foreign * exchange_rate / 10000, 4) if exchange_rate > 0 else round(total_foreign / 10000, 4),
        }
        for col_idx, label in month_cols:
            val = _safe_float(row[col_idx]) if col_idx < len(row) else 0
            rec[label] = val

        results.append(rec)

    return pd.DataFrame(results) if results else None


def _manual_sum_months(row, header_row2, total_col_start):
    """手动汇总各月的应收/已收"""
    total_recv = 0
    total_paid = 0
    for i in range(1, total_col_start):
        val = _safe_float(row[i]) if i < len(row) else 0
        if val and i < len(header_row2):
            sub_header = str(header_row2[i]) if header_row2[i] else ''
            if '应收金额' in sub_header:
                total_recv += val
            elif '已收金额' in sub_header:
                total_paid += val
    return total_recv, total_paid, total_recv - total_paid


def _parse_payable(fpath: str) -> Optional[pd.DataFrame]:
    """
    解析应付账款明细表。支持两种格式：
    格式1 (矩阵): 客户名称 + 各月(应付/已付/余额)
    格式2 (简单): 供应商名称 + 各月金额 + 总计
    """
    data = _read_all_rows(fpath)
    if len(data) < 3:
        return None

    header_texts = ' '.join([str(c) for c in (data[0] + (data[1] if len(data) > 1 else [])) if c])

    if '应付金额' in header_texts or '已付金额' in header_texts:
        return _parse_payable_matrix(data)
    else:
        return _parse_payable_simple(data)


def _parse_payable_matrix(data: list) -> Optional[pd.DataFrame]:
    """解析矩阵格式应付"""
    if len(data) < 4:
        return None

    header_row2 = data[1]
    total_cols = len(header_row2)

    results = []
    for row_idx in range(3, len(data)):
        row = data[row_idx]
        if not row or not row[0] or str(row[0]).strip() == '':
            continue

        supplier = str(row[0]).strip()
        total_payable = 0
        total_paid = 0
        for i in range(1, total_cols):
            val = _safe_float(row[i]) if i < len(row) else 0
            if val and i < len(header_row2):
                sub_header = str(header_row2[i]) if header_row2[i] else ''
                if '应付金额' in sub_header:
                    total_payable += val
                elif '已付金额' in sub_header:
                    total_paid += val

        balance = total_payable - total_paid
        results.append({
            '供应商': supplier,
            '应付余额': balance / 10000,
        })

    return pd.DataFrame(results)


def _parse_payable_simple(data: list) -> Optional[pd.DataFrame]:
    """
    解析简单格式应付（如和钛的格式）。
    格式: 标题行 / 表头行(供应商名称 + 月份 + 总计) / 数据行
    """
    import re

    # 找到标题行（通常第0行是标题如 "2025.12月-2026.3月供应商对账明细表"）
    # 找到真正的表头行（含"供应商"字样）
    title_row_idx = 0
    header_row_idx = None
    for i, row in enumerate(data[:5]):
        row_text = [str(c).strip() for c in row if c]
        if any('供应商' in t for t in row_text):
            header_row_idx = i
            break

    # 如果没找到独立表头行，表头可能在第1个数据行中
    if header_row_idx is None:
        # 用第0行作为标题，第一个数据行检查是否是隐式表头
        header_row_idx = 0

    headers_raw = data[header_row_idx]
    headers = [str(c).strip() if c else '' for c in headers_raw]

    # 找供应商列和总计列
    name_col = 0
    total_col = None
    month_cols = []  # (col_index, month_label)
    for i, h in enumerate(headers):
        if '供应商' in h or '名称' in h:
            name_col = i
        elif '总计' in h or '合计' in h:
            total_col = i
        else:
            # 检查是否为日期列 (如 "2025-12-01" 或 "2025.12")
            if isinstance(headers_raw[i], datetime):
                dt = headers_raw[i]
                label = dt.strftime('%Y-%m')
                month_cols.append((i, label))
            else:
                m_month = re.search(r'(\d{4})[\.\-](\d{1,2})', h)
                if m_month:
                    label = f"{m_month.group(1)}-{int(m_month.group(2)):02d}"
                    month_cols.append((i, label))

    results = []
    for row_idx in range(header_row_idx + 1, len(data)):
        row = data[row_idx]
        if not row:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) else ''
        if not name or name in ('合计', '总计', '小计', '', 'nan', 'None'):
            continue
        # 跳过重复的表头行（第一行数据可能是 "供应商名称"）
        if name in ('供应商名称', '供应商', '名称'):
            continue

        total = _safe_float(row[total_col]) if total_col and total_col < len(row) else 0
        if total_col is None:
            total = sum(_safe_float(row[i]) for i in range(name_col + 1, len(row)))

        rec = {
            '供应商': name,
            '合计(万元)': round(total / 10000, 4),
        }
        for col_idx, label in month_cols:
            val = _safe_float(row[col_idx]) if col_idx < len(row) else 0
            rec[label] = round(val / 10000, 4)

        results.append(rec)

    return pd.DataFrame(results) if results else None


def _safe_float(val) -> float:
    """安全转换为 float"""
    if val is None or val == '':
        return 0.0
    try:
        if isinstance(val, str):
            if val.startswith('='):
                return 0.0
            val = val.replace(',', '').replace('，', '')
        return float(val)
    except (ValueError, TypeError):
        return 0.0
