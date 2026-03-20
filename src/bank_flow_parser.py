# -*- coding: utf-8 -*-
"""对公流水解析模块 — 支持 .xls / .xlsx / .csv / .pdf（银行导出流水），自动去重合并"""

import pandas as pd
import xlrd
import openpyxl
from datetime import datetime, timedelta
from typing import List, Optional
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning)

# 标准列名
STANDARD_COLUMNS = [
    '交易时间', '收入金额', '支出金额', '账户余额',
    '对方账号', '对方户名', '对方开户行', '交易用途', '摘要'
]


def _to_float(s: str) -> float:
    """将金额字符串（含逗号、括号负数）转为浮点数，无效则返回0.0"""
    s = re.sub(r'[,，\s]', '', str(s))
    s = s.replace('(', '-').replace('（', '-').replace(')', '').replace('）', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def parse(file_paths: List[str]) -> pd.DataFrame:
    """
    解析多个流水文件并合并去重。

    Args:
        file_paths: 流水文件路径列表

    Returns:
        合并去重后的 DataFrame，列名标准化
    """
    all_dfs = []
    for fpath in file_paths:
        try:
            df = _parse_single_file(fpath)
            if df is not None and len(df) > 0:
                all_dfs.append(df)
        except Exception as e:
            print(f"[WARN] 解析流水文件失败: {fpath}, 错误: {e}")

    if not all_dfs:
        return pd.DataFrame(columns=STANDARD_COLUMNS)

    merged = pd.concat(all_dfs, ignore_index=True)

    # 去重：基于 交易时间 + 收入金额 + 支出金额 + 对方户名
    merged = _deduplicate(merged)

    # 按交易时间排序
    merged = merged.sort_values('交易时间').reset_index(drop=True)

    return merged


def _parse_single_file(fpath: str) -> Optional[pd.DataFrame]:
    """解析单个流水文件"""
    ext = fpath.lower()
    if ext.endswith('.xlsx'):
        return _parse_xlsx(fpath)
    elif ext.endswith('.xls'):
        return _parse_xls(fpath)
    elif ext.endswith('.csv'):
        return _parse_csv(fpath)
    elif ext.endswith('.pdf'):
        return _parse_pdf(fpath)
    return None


# ---------------------------------------------------------------------------
# PDF 流水解析（农商行/部分银行打印导出格式）
# ---------------------------------------------------------------------------
def _parse_pdf(fpath: str) -> Optional[pd.DataFrame]:
    """
    解析银行流水 PDF（pdfplumber 表格抽取）。

    支持格式：
    - 珠海农商行账户明细查询 PDF
    - 其他含标准列名（交易日期/收入/支出/余额/对方）的银行流水 PDF
    """
    try:
        import pdfplumber
    except ImportError:
        print("[WARN] 未安装 pdfplumber，无法解析 PDF 流水")
        return None

    all_rows = []
    header_cols = None  # 已确定的列名映射: {col_index: 标准列名}

    # 农商行列名 → 标准列名 映射关键词
    _PDF_COL_MAP = {
        "交易日期": "交易时间",
        "收入":     "收入金额",
        "支出":     "支出金额",
        "账户余额": "账户余额",
        "对方账号": "对方账号",
        "对方户名": "对方户名",
        "对方行名": "对方开户行",
        "附言":     "交易用途",
        "交易类型": "摘要",
    }

    try:
        with pdfplumber.open(fpath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row is None:
                            continue
                        vals = [str(v).replace("\n", " ").strip() if v else "" for v in row]

                        # 识别表头行
                        if header_cols is None:
                            if "交易日期" in vals or "收入" in vals:
                                header_cols = {}
                                for i, v in enumerate(vals):
                                    for pdf_col, std_col in _PDF_COL_MAP.items():
                                        if pdf_col in v:
                                            header_cols[i] = std_col
                                            break
                            continue

                        # 数据行：第一列应为流水号（纯数字或含字母）
                        if not vals[0] or not re.match(r'^[\w]+$', vals[0].strip()):
                            continue

                        record = {col: "" for col in STANDARD_COLUMNS}
                        for i, std_col in header_cols.items():
                            if i < len(vals):
                                record[std_col] = vals[i]
                        all_rows.append(record)

    except Exception as e:
        print(f"[WARN] PDF流水解析失败: {e}")
        return None

    if not all_rows:
        print(f"  [WARN] PDF中未找到流水数据: {fpath}")
        return None

    df = pd.DataFrame(all_rows, columns=STANDARD_COLUMNS)

    # 清洗金额列
    for col in ("收入金额", "支出金额", "账户余额"):
        df[col] = df[col].apply(lambda x: _to_float(str(x)))

    # 解析交易时间
    def _parse_dt(s: str):
        s = s.strip()
        # 尝试包含时分秒的格式 (YYYY-MM-DD HH:MM:SS), 19 chars
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                return datetime.strptime(s[:19], fmt)
            except (ValueError, IndexError):
                pass
        # 降级：只解析日期部分, 10 chars
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except (ValueError, IndexError):
            pass
        return pd.NaT

    df["交易时间"] = df["交易时间"].apply(_parse_dt)
    df = df.dropna(subset=["交易时间"])
    print(f"  [PDF] 解析到 {len(df)} 条流水记录")
    return df


def _parse_xlsx(fpath: str) -> Optional[pd.DataFrame]:
    """解析 .xlsx 格式流水文件"""
    wb = openpyxl.load_workbook(fpath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 跳过非流水 sheet（如收入/支出汇总透视表）
        if sheet_name in ('收入', '支出') or '汇总' in sheet_name:
            continue

        # 查找表头行 — 支持多种格式
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True), 1):
            row_text = [str(c) if c else '' for c in row]
            if _is_header_row(row_text):
                header_row = row_idx
                break

        if header_row is None:
            continue

        # 读取数据
        rows = []
        headers = [str(c.value) if c.value else '' for c in ws[header_row]]

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if row[0] is None:
                continue
            rows.append(list(row))

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
        df = _standardize_columns(df)
        if df is not None:
            return df

    return None


def _parse_csv(fpath: str) -> Optional[pd.DataFrame]:
    """解析 .csv 格式流水文件（支持 GBK / UTF-8 / UTF-8-sig 编码）"""
    # 尝试多种编码
    for encoding in ('gbk', 'utf-8-sig', 'utf-8', 'gb2312', 'gb18030'):
        try:
            df_raw = pd.read_csv(fpath, encoding=encoding, dtype=str, header=None, skip_blank_lines=True)
            if df_raw.empty:
                continue

            # 查找表头行（在前10行中）
            header_row = None
            for row_idx in range(min(10, len(df_raw))):
                row_vals = [str(v).strip() for v in df_raw.iloc[row_idx].tolist()]
                if _is_header_row(row_vals):
                    header_row = row_idx
                    break

            if header_row is None:
                continue

            headers = [str(v).strip() for v in df_raw.iloc[header_row].tolist()]
            rows = df_raw.iloc[header_row + 1:].values.tolist()
            # 去除全空行
            rows = [r for r in rows if any(str(v).strip() not in ('', 'nan', 'None') for v in r)]

            if not rows:
                continue

            df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
            df = _standardize_columns(df)
            if df is not None:
                return df

        except (UnicodeDecodeError, Exception):
            continue

    return None


def _parse_xls(fpath: str) -> Optional[pd.DataFrame]:
    """解析 .xls 格式流水文件"""
    wb = xlrd.open_workbook(fpath)

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)

        # 跳过汇总类 sheet
        if sheet_name in ('收入', '支出') or '汇总' in sheet_name:
            continue

        # 查找表头行
        header_row = None
        for row_idx in range(min(10, ws.nrows)):
            row_vals = [str(ws.cell_value(row_idx, col)) for col in range(ws.ncols)]
            if _is_header_row(row_vals):
                header_row = row_idx
                break

        if header_row is None:
            continue

        headers = [str(ws.cell_value(header_row, col)) for col in range(ws.ncols)]

        rows = []
        for row_idx in range(header_row + 1, ws.nrows):
            row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
            if row_vals[0] is None or row_vals[0] == '':
                continue
            rows.append(row_vals)

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
        df = _standardize_columns(df)
        if df is not None:
            return df

    return None


def _is_header_row(row_text: list) -> bool:
    """判断一行是否为表头行，兼容多种格式"""
    text_set = set(row_text)
    # 格式1: 标准流水 (交易时间/收入金额/支出金额)
    if '交易时间' in text_set:
        return True
    # 格式2: 对账单 (序号/交易日期/借贷标记/对方户名)
    if '序号' in text_set and ('交易日期' in text_set or '对方户名' in text_set):
        return True
    # 格式3: 手工日记账 (日期/付款方/收款方/摘要)
    if '日期' in text_set and ('付款方' in text_set or '收款方' in text_set):
        return True
    return False


def _standardize_columns(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """标准化列名并转换数据类型，支持多种流水格式"""

    col_names = [str(c).strip() for c in df.columns]

    # 检测格式类型
    if '借贷标记' in col_names or '借贷标志' in col_names:
        return _standardize_bank_statement(df)
    elif '付款方' in col_names or '收款方' in col_names:
        return _standardize_journal(df)
    else:
        return _standardize_standard(df)


def _standardize_standard(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """标准化标准流水格式（交易时间/收入金额/支出金额）"""
    col_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if '交易时间' in col_str or ('日期' in col_str and '时间' not in col_str and '交易时间' not in col_map.values()):
            if '交易时间' not in col_map.values():
                col_map[col] = '交易时间'
        elif col_str == '时间':
            pass  # 跳过纯时间列
        elif '收入' in col_str:
            col_map[col] = '收入金额'
        elif '支出' in col_str:
            col_map[col] = '支出金额'
        elif '余额' in col_str:
            col_map[col] = '账户余额'
        elif '对方账号' in col_str or '对方帐号' in col_str:
            col_map[col] = '对方账号'
        elif '对方户名' in col_str:
            col_map[col] = '对方户名'
        elif '开户行' in col_str:
            col_map[col] = '对方开户行'
        elif '用途' in col_str:
            col_map[col] = '交易用途'
        elif '摘要' in col_str:
            col_map[col] = '摘要'

    if '交易时间' not in col_map.values():
        return None

    df = df.rename(columns=col_map)
    return _finalize_df(df)


def _standardize_bank_statement(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    标准化对账单格式（借贷标记 + 交易金额，或借方发生额/贷方发生额分列格式）。
    格式1: 借贷标记 + 交易金额（中行、建行等）
    格式2: 借方发生额 + 贷方发生额（工行 HISTORYDETAIL 格式）
    """
    col_map = {}
    borrow_lend_col = None
    amount_col = None
    debit_col = None   # 借方发生额（支出）
    credit_col = None  # 贷方发生额（收入）

    for col in df.columns:
        col_str = str(col).strip()
        if '交易日期' in col_str or col_str == '日期':
            col_map[col] = '交易时间'
        elif col_str == '交易时间':
            # 如果已经有 交易日期 列被标记为主日期，则此列是纯时刻列(辅助)
            # 否则（如工行格式只有交易时间），此列就是主日期列
            if any('交易日期' in str(c) or str(c).strip() == '日期' for c in df.columns):
                col_map[col] = '_交易时刻'
            else:
                col_map[col] = '交易时间'
        elif '借贷标' in col_str:
            borrow_lend_col = col
        elif col_str == '交易金额':
            amount_col = col
        elif col_str in ('借方发生额', '借方金额', '借方'):
            debit_col = col
        elif col_str in ('贷方发生额', '贷方金额', '贷方'):
            credit_col = col
        elif '余额' in col_str:
            col_map[col] = '账户余额'
        elif '对方账号' in col_str or '对方帐号' in col_str:
            col_map[col] = '对方账号'
        elif '对方户名' in col_str:
            col_map[col] = '对方户名'
        elif '摘要' in col_str and '代码' not in col_str:
            col_map[col] = '摘要'
        elif '用途' in col_str:
            col_map[col] = '交易用途'

    if '交易时间' not in col_map.values():
        return None

    df = df.rename(columns=col_map)

    # 格式2：借方发生额 + 贷方发生额 分列（工行格式）
    if debit_col is not None and credit_col is not None:
        df['支出金额'] = df[debit_col].apply(_parse_amount)
        df['收入金额'] = df[credit_col].apply(_parse_amount)
    # 格式1：借贷标记 + 交易金额
    elif borrow_lend_col is not None and amount_col is not None:
        df['收入金额'] = df.apply(
            lambda r: _parse_amount(r[amount_col]) if str(r[borrow_lend_col]).strip() in ('贷', 'C', 'CR') else 0.0,
            axis=1
        )
        df['支出金额'] = df.apply(
            lambda r: _parse_amount(r[amount_col]) if str(r[borrow_lend_col]).strip() in ('借', 'D', 'DR') else 0.0,
            axis=1
        )

    # 摘要 → 交易用途
    if '摘要' in df.columns and '交易用途' not in df.columns:
        df['交易用途'] = df['摘要']

    return _finalize_df(df)


def _standardize_journal(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    标准化手工日记账格式。
    列: 日期/付款方/收款方/摘要/本期收入/本期支出/余额/备注
    """
    col_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str == '日期':
            col_map[col] = '交易时间'
        elif '收入' in col_str:
            col_map[col] = '收入金额'
        elif '支出' in col_str:
            col_map[col] = '支出金额'
        elif '余额' in col_str:
            col_map[col] = '账户余额'
        elif '摘要' in col_str:
            col_map[col] = '交易用途'

    if '交易时间' not in col_map.values():
        return None

    df = df.rename(columns=col_map)

    # 临时保留付款方和收款方，从中推断对方户名
    pay_col = None
    recv_col = None
    for col in df.columns:
        if '付款方' in str(col):
            pay_col = col
        if '收款方' in str(col):
            recv_col = col

    if pay_col and recv_col:
        # 收入时对方是付款方，支出时对方是收款方
        df['对方户名'] = df.apply(
            lambda r: str(r[pay_col]).strip() if _parse_amount(r.get('收入金额', 0)) > 0
                      else str(r[recv_col]).strip(), axis=1
        )

    return _finalize_df(df)


def _finalize_df(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """最终标准化：只保留标准列，转换类型"""
    # 只保留标准列（有的列可能缺失）
    existing_cols = [c for c in STANDARD_COLUMNS if c in df.columns]
    df = df[existing_cols].copy()

    # 补齐缺失列
    for col in STANDARD_COLUMNS:
        if col not in df.columns:
            df[col] = ''

    # 转换交易时间
    df['交易时间'] = df['交易时间'].apply(_parse_datetime)

    # 转换金额列
    for col in ['收入金额', '支出金额', '账户余额']:
        df[col] = df[col].apply(_parse_amount)

    # 转换文本列
    for col in ['对方账号', '对方户名', '对方开户行', '交易用途', '摘要']:
        df[col] = df[col].apply(lambda x: str(x).strip() if pd.notna(x) and x != '' else '')

    # 移除交易时间为空的行
    df = df.dropna(subset=['交易时间'])

    return df


def _parse_datetime(val) -> Optional[datetime]:
    """
    解析日期时间值。支持：
    - 字符串格式: '2025-01-02 15:10:20'
    - Excel 序列号: 45659.632175...
    - datetime 对象
    """
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return None

    if isinstance(val, datetime):
        return val

    if isinstance(val, (int, float)):
        try:
            # Excel serial number (1900-based)
            # Excel 序列号: 1 = 1900-01-01
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except (ValueError, OverflowError):
            return None

    val_str = str(val).strip()
    # 尝试多种日期格式
    for fmt in [
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y%m%d',
    ]:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue

    return None


def _parse_amount(val) -> float:
    """解析金额值，处理空值、字符串、数字"""
    if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
        return 0.0
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return 0.0
        return float(val)
    # 去掉逗号等格式
    val_str = str(val).replace(',', '').replace('，', '').strip()
    try:
        return float(val_str)
    except ValueError:
        return 0.0


def _deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    """
    基于（交易时间, 收入金额, 支出金额, 对方户名）去重。
    保留第一条记录。
    """
    # 创建去重 key
    df['_dedup_key'] = (
        df['交易时间'].astype(str) + '|' +
        df['收入金额'].astype(str) + '|' +
        df['支出金额'].astype(str) + '|' +
        df['对方户名'].astype(str)
    )
    df = df.drop_duplicates(subset='_dedup_key', keep='first')
    df = df.drop(columns=['_dedup_key'])
    return df
