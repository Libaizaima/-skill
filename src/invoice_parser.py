# -*- coding: utf-8 -*-
"""发票解析模块 — 解析进项票/销项票统计表"""

import re
import pandas as pd
import xlrd
import openpyxl
from datetime import datetime, timedelta
from typing import List, Tuple, Optional
import warnings

warnings.filterwarnings('ignore', category=UserWarning)


def parse(invoices_in_paths: List[str], invoices_out_paths: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    解析进项票和销项票统计表。

    Args:
        invoices_in_paths: 进项票统计文件路径列表
        invoices_out_paths: 销项票统计文件路径列表

    Returns:
        (in_df, out_df): 进项票和销项票 DataFrame
        每个 DataFrame 列: ['开票日期', '销方名称', '购买方名称', '金额', '税额', '价税合计', '发票状态']
    """
    in_dfs = []
    for fpath in invoices_in_paths:
        try:
            df = _parse_invoice_file(fpath, direction='in')
            if df is not None and len(df) > 0:
                in_dfs.append(df)
        except Exception as e:
            print(f"[WARN] 解析进项票文件失败: {fpath}, 错误: {e}")

    out_dfs = []
    for fpath in invoices_out_paths:
        try:
            df = _parse_invoice_file(fpath, direction='out')
            if df is not None and len(df) > 0:
                out_dfs.append(df)
        except Exception as e:
            print(f"[WARN] 解析销项票文件失败: {fpath}, 错误: {e}")

    in_df = pd.concat(in_dfs, ignore_index=True) if in_dfs else pd.DataFrame()
    out_df = pd.concat(out_dfs, ignore_index=True) if out_dfs else pd.DataFrame()

    # 只保留发票状态为"正常"的记录
    if len(in_df) > 0 and '发票状态' in in_df.columns:
        in_df = in_df[in_df['发票状态'] == '正常'].copy()
    if len(out_df) > 0 and '发票状态' in out_df.columns:
        out_df = out_df[out_df['发票状态'] == '正常'].copy()

    return in_df, out_df


def _parse_invoice_file(fpath: str, direction: str = 'in') -> Optional[pd.DataFrame]:
    """解析单个发票统计文件"""
    if fpath.lower().endswith('.xlsx'):
        return _parse_invoice_xlsx(fpath, direction)
    elif fpath.lower().endswith('.xls'):
        return _parse_invoice_xls(fpath, direction)
    return None


def _parse_invoice_xlsx(fpath: str, direction: str) -> Optional[pd.DataFrame]:
    """解析 .xlsx 格式发票文件"""
    wb = openpyxl.load_workbook(fpath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 找到包含 "序号" 或 "发票" 的表头行
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True), 1):
            row_text = [str(c) if c else '' for c in row]
            if '序号' in row_text and ('开票日期' in row_text or '价税合计' in row_text):
                header_row = row_idx
                break

        if header_row is None:
            continue

        headers = [str(c.value) if c.value else '' for c in ws[header_row]]
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if row[0] is None or str(row[0]).strip() == '':
                continue
            rows.append(list(row))

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
        return _standardize_invoice_columns(df, direction)

    return None


def _parse_invoice_xls(fpath: str, direction: str) -> Optional[pd.DataFrame]:
    """解析 .xls 格式发票文件"""
    wb = xlrd.open_workbook(fpath)

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)

        # 跳过汇总类 sheet
        if ws.nrows < 2:
            continue

        # 找表头
        header_row = None
        for row_idx in range(min(10, ws.nrows)):
            row_vals = [str(ws.cell_value(row_idx, col)) for col in range(ws.ncols)]
            if '序号' in row_vals and ('开票日期' in row_vals or '价税合计' in row_vals):
                header_row = row_idx
                break

        if header_row is None:
            continue

        headers = [str(ws.cell_value(header_row, col)) for col in range(ws.ncols)]
        rows = []
        for row_idx in range(header_row + 1, ws.nrows):
            row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
            if row_vals[0] is None or str(row_vals[0]).strip() == '':
                continue
            rows.append(row_vals)

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
        return _standardize_invoice_columns(df, direction)

    return None


def _standardize_invoice_columns(df: pd.DataFrame, direction: str) -> Optional[pd.DataFrame]:
    """标准化发票 DataFrame"""
    col_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if '开票日期' in col_str:
            col_map[col] = '开票日期'
        elif '销方名称' in col_str:
            col_map[col] = '销方名称'
        elif '购买方名称' in col_str or '购方名称' in col_str:
            col_map[col] = '购买方名称'
        elif col_str == '金额':
            col_map[col] = '金额'
        elif col_str == '税额':
            col_map[col] = '税额'
        elif '价税合计' in col_str:
            col_map[col] = '价税合计'
        elif '发票状态' in col_str:
            col_map[col] = '发票状态'

    df = df.rename(columns=col_map)

    # 确保关键列存在
    if '价税合计' not in df.columns:
        return None

    # 转换开票日期
    if '开票日期' in df.columns:
        df['开票日期'] = df['开票日期'].apply(_parse_invoice_date)

    # 转换金额列
    for col in ['金额', '税额', '价税合计']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

    # 提取币种信息（从备注字段）
    df = _extract_currency_info(df)

    # 添加方向标记
    df['方向'] = direction

    return df


def _parse_invoice_date(val) -> Optional[datetime]:
    """解析发票日期"""
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return None

    if isinstance(val, datetime):
        return val

    if isinstance(val, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except (ValueError, OverflowError):
            return None

    val_str = str(val).strip()
    for fmt in [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d',
    ]:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue

    return None


# ── 币种信息提取 ──

_CURRENCY_MAP = {
    '美元': 'USD', '美金': 'USD', 'USD': 'USD',
    '欧元': 'EUR', 'EUR': 'EUR',
    '英镑': 'GBP', 'GBP': 'GBP',
    '日元': 'JPY', '日币': 'JPY', 'JPY': 'JPY',
    '港币': 'HKD', '港元': 'HKD', 'HKD': 'HKD',
    '人民币': 'CNY', 'CNY': 'CNY', 'RMB': 'CNY',
}


def _extract_currency_info(df: pd.DataFrame) -> pd.DataFrame:
    """
    从备注字段提取币种、外币金额、汇率。
    出口发票备注格式: "币别：美元、外币出口销售额：121770、汇率：6.9678"
    """
    df['币种'] = 'CNY'
    df['外币金额'] = 0.0
    df['汇率'] = 1.0

    if '备注' not in df.columns:
        return df

    for idx, row in df.iterrows():
        remark = str(row.get('备注', ''))
        if not remark or remark == 'nan':
            continue

        # 提取币别
        m = re.search(r'币别[：:]\s*(\S+)', remark)
        if m:
            currency_raw = m.group(1).strip('、，, ')
            code = _CURRENCY_MAP.get(currency_raw, currency_raw.upper())
            if code != 'CNY':
                df.at[idx, '币种'] = code

                # 提取外币金额
                m_amt = re.search(r'外币.*?(?:销售额|金额)[：:]?\s*([\d,.]+)', remark)
                if m_amt:
                    try:
                        df.at[idx, '外币金额'] = float(m_amt.group(1).replace(',', ''))
                    except ValueError:
                        pass

                # 提取汇率
                m_rate = re.search(r'汇率[：:]?\s*([\d.]+)', remark)
                if m_rate:
                    try:
                        df.at[idx, '汇率'] = float(m_rate.group(1))
                    except ValueError:
                        pass

    return df
